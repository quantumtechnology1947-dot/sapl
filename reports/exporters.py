"""
Report Export Utilities
Handles PDF and Excel export for all reports
"""

from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from datetime import datetime


class ExcelExporter:
    """
    Export reports to Excel format using openpyxl.
    
    TODO: Install openpyxl: pip install openpyxl
    """
    
    @staticmethod
    def export_boughtout_design(queryset, filters=None):
        """Export Boughtout Design Report to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Boughtout Design"
            
            # Add title
            ws['A1'] = 'Boughtout Design Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            
            # Add filters info
            row = 4
            if filters:
                ws[f'A{row}'] = 'Filters Applied:'
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                for key, value in filters.items():
                    if value:
                        ws[f'A{row}'] = f'{key}: {value}'
                        row += 1
                row += 1
            
            # Headers
            headers = ['SN', 'Item Code', 'Description', 'Category', 'Unit', 
                      'Qty Required', 'Qty Ordered', 'Qty Received', 'Pending', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            row += 1
            for idx, item in enumerate(queryset, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item.itemcode)
                ws.cell(row=row, column=3, value=item.itemdescription)
                ws.cell(row=row, column=4, value=item.category.name if item.category else '-')
                ws.cell(row=row, column=5, value=item.unit.name if item.unit else '-')
                ws.cell(row=row, column=6, value='--')
                ws.cell(row=row, column=7, value='--')
                ws.cell(row=row, column=8, value='--')
                ws.cell(row=row, column=9, value='--')
                ws.cell(row=row, column=10, value='Pending')
                row += 1
            
            # Adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except ImportError:
            # openpyxl not installed
            return None
    
    @staticmethod
    def export_boughtout_vendor(queryset, filters=None):
        """Export Boughtout Vendor Report to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Boughtout Vendor"
            
            # Title
            ws['A1'] = 'Boughtout Vendor Report'
            ws['A1'].font = Font(size=16, bold=True)
            ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
            
            # Headers
            row = 4
            headers = ['SN', 'Supplier', 'Item Code', 'Description', 'Quantity', 'Rate', 'Amount', 'PO No']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Data
            row += 1
            for idx, item in enumerate(queryset, 1):
                ws.cell(row=row, column=1, value=idx)
                ws.cell(row=row, column=2, value=item.po.supplier.name if item.po and item.po.supplier else '-')
                ws.cell(row=row, column=3, value=item.item.itemcode if item.item else '-')
                ws.cell(row=row, column=4, value=item.item.itemdescription if item.item else '-')
                ws.cell(row=row, column=5, value=float(item.quantity) if item.quantity else 0)
                ws.cell(row=row, column=6, value=float(item.rate) if item.rate else 0)
                ws.cell(row=row, column=7, value=float(item.amount) if item.amount else 0)
                ws.cell(row=row, column=8, value=item.po.po_no if item.po else '-')
                row += 1
            
            # Adjust widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
            
        except ImportError:
            return None


class PDFExporter:
    """
    Export reports to PDF format using ReportLab.
    
    TODO: Install reportlab: pip install reportlab
    """
    
    @staticmethod
    def export_report(title, headers, data, filters=None):
        """Generic PDF export for reports."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#1f2937'),
                spaceAfter=30,
            )
            elements.append(Paragraph(title, title_style))
            
            # Generated date
            date_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            elements.append(Paragraph(date_text, styles['Normal']))
            elements.append(Spacer(1, 0.2*inch))
            
            # Filters
            if filters:
                elements.append(Paragraph("<b>Filters Applied:</b>", styles['Normal']))
                for key, value in filters.items():
                    if value:
                        elements.append(Paragraph(f"{key}: {value}", styles['Normal']))
                elements.append(Spacer(1, 0.3*inch))
            
            # Table
            table_data = [headers] + data
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            
            doc.build(elements)
            output.seek(0)
            return output
            
        except ImportError:
            return None


def create_excel_response(output, filename):
    """Create HTTP response for Excel file."""
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def create_pdf_response(output, filename):
    """Create HTTP response for PDF file."""
    response = HttpResponse(output.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
